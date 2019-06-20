from openpyxl import load_workbook
from collections import namedtuple
from scripts.loggerhandler import do_logger
from scripts.confighandler import do_config
from scripts.constants import CASE_DATA_DIR
from scripts.ReHandler import ReText


class ExcelHandler(object):
    """
    封装读写excel的操作
    """
    def __init__(self, filename, sheetname='', max_col=do_config('excel', 'max_column')):
        """
        初始化
        :param filename: excel文件路径名
        :param sheetname: 需要定位的表单名
        :param max_col:  读取的最大列数（扣除待写的列）
        """
        self.filename = filename  # 从常量文件中读取excel绝对路径
        self.sheetname = sheetname
        self.data_list = []
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[sheetname] if sheetname != '' else self.wb.active
        self.max_col = max_col
        self.data_header = tuple(self.ws.iter_rows(max_row=1, max_col=self.max_col, values_only=True))[0]
        self.Cases = namedtuple("Cases", self.data_header)

    def write_excel(self, row, *data_tuple):
        """
        写入excel
        :param row: 单元格的行号 int
        :param data_tuple: 待写入的值
        """
        other_wb = load_workbook(self.filename)
        other_ws = other_wb[self.sheetname]
        if isinstance(row, int) and (1 < row <= other_ws.max_row):
            try:
                columns = do_config('excel', 'columns', iseval=True)  # 读取配置文件中单元格的列号，类型可以为int，list，tuple，''
            except SyntaxError:  # columns选项值为空时或者类型不正确时，默认写入列号为max_col的后两个
                if len(data_tuple) != 2:
                    do_logger.error('写入的列数与值的个数不匹配，写入失败！')
                    raise IndexError('tuple index out of range')
                other_ws.cell(row, self.max_col+1, value=data_tuple[0])
                other_ws.cell(row, self.max_col+2, value=data_tuple[1])
                other_wb.save(self.filename)
            else:
                if isinstance(columns, (list, tuple)):
                    sort_columns = list(set(columns))
                    sort_columns.sort()
                    if len(data_tuple) != len(sort_columns):
                        do_logger.error('写入的列数与值的个数不匹配，写入失败！')
                        raise IndexError('tuple index out of range')
                    for i in sort_columns:
                        other_ws.cell(row, i, value=data_tuple[sort_columns.index(i)])
                    other_wb.save(self.filename)
                elif isinstance(columns, int):
                    other_ws.cell(row, columns, value=data_tuple[0])
                    other_wb.save(self.filename)
        else:
            do_logger.error('传入的行号错误，行号必须是大于1的整数！写入失败！')
            
    def read_excel(self):
        """
        读取excel
        :return:嵌套命名元组的列表
        """
        for tuple_data in self.ws.iter_rows(min_row=2, max_col=self.max_col, values_only=True):
            self.data_list.append(self.Cases(*tuple_data))
        return self.data_list
    
        
# sheetnames = do_config('excel', 'sheet_name', iseval=True)    # 处理多个表单的情况下，实例化多个ExcelHandler对象
# register_excel = ExcelHandler(sheetname=sheetnames[0], max_col=do_config('excel', 'max_column'))
# login_excel = ExcelHandler(sheetname=sheetnames[1], max_col=do_config('excel', 'max_column'))

if __name__ == '__main__':
    # a = register_excel.read_excel()
    # b = ReText().register_parametrization(a[0].data)
    # print(eval(b))
    pass
